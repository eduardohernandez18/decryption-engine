# pip install openpyxl

import openpyxl
import re

dic = {
    ' ': '404741404142',
    '.': '4017',
    '-': '4016',
    '/': '4014',
    '(': '404A',
    ')': '404B',
    '+': '4010',
    'A': '4643',
    'B': '4640',
    'C': '4641',
    'D': '4646',
    'E': '4647',
    'F': '4644',
    'G': '4645',
    'H': '464A',
    'I': '464B',
    'J': '4613',
    'K': '4610',
    'L': '4611',
    'M': '4616',
    'N': '4617',
    'O': '4614',
    'P': '4742',
    'Q': '4743',
    'R': '4740',
    'S': '4741',
    'T': '4746',
    'U': '4747',
    'V': '4744',
    'W': '4745',
    'X': '474A',
    'Y': '474B',
    'Z': '4713',
    '0': '4142',
    '1': '4143',
    '2': '4140',
    '3': '4141',
    '4': '4146',
    '5': '4147',
    '6': '4144',
    '7': '4145',
    '8': '414A',
    '9': '414B'
}

wb = openpyxl.load_workbook('data2.xlsx')
ws = wb.active
json_data = []

for row in ws.iter_rows(min_row=2, values_only=True):
    encoded, decoded = row

    if len(encoded) % 4 != 0:
        print(f"row {ws.max_row} is an unknown format!")
    else:
        expected_decoded = ''
        i = 0
        while i < len(encoded):
            chunk1 = encoded[i:i+12] if i + 12 <= len(encoded) else None
            chunk2 = encoded[i:i+4]
            k = None
            
            if chunk1 and (k := next((key for key, value in dic.items() if value == chunk1), None)):
                expected_decoded += k
                i += 12
            elif k := next((key for key, value in dic.items() if value == chunk2), None):
                expected_decoded += k
                i += 4
            else:
                print(f"not found in dic: {chunk2}")
                break
          
        if expected_decoded != decoded:
            print(f"row {ws.max_row} wrong! '{encoded}' to '{decoded}', expecting: '{expected_decoded}'")

wb.close()
